#!/usr/bin/env python3
"""
NISTify 800-171 R2 - NIST SP 800-171 Compliance Scanner and Reporter
Windows Compatible Version (No Emojis)
Scans networks and endpoints for compliance with NIST SP 800-171 Rev 2
Generates compliance reports in multiple formats and POA&M documents
"""

import os
import sys
import json
import xml.etree.ElementTree as ET
from xml.dom import minidom
import socket
import subprocess
import platform
import datetime
import argparse
import logging
from pathlib import Path
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, asdict
from typing import List, Dict, Optional, Tuple
import ipaddress

# Third-party imports
try:
    import nmap
    import pandas as pd
    from jinja2 import Template
    import pdfkit
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    import requests
    import networkx as nx
    import matplotlib.pyplot as plt
    import matplotlib.patches as mpatches
    from matplotlib.patches import FancyBboxPatch
    import numpy as np
except ImportError as e:
    print(f"Missing required package: {e}")
    print("Install with: pip install python-nmap pandas jinja2 pdfkit openpyxl requests networkx matplotlib")
    sys.exit(1)

def print_banner():
    """Print ASCII art banner for NISTify 800-171 R2"""
    banner = """
    ╔══════════════════════════════════════════════════════════════════════════════════════════╗
    ║                                                                                          ║
    ║    ███╗   ██╗██╗███████╗████████╗██╗███████╗██╗   ██╗                                    ║
    ║    ████╗  ██║██║██╔════╝╚══██╔══╝██║██╔════╝╚██╗ ██╔╝                                    ║
    ║    ██╔██╗ ██║██║███████╗   ██║   ██║█████╗   ╚████╔╝                                     ║
    ║    ██║╚██╗██║██║╚════██║   ██║   ██║██╔══╝    ╚██╔╝                                      ║
    ║    ██║ ╚████║██║███████║   ██║   ██║██║        ██║                                       ║
    ║    ╚═╝  ╚═══╝╚═╝╚══════╝   ╚═╝   ╚═╝╚═╝        ╚═╝                                       ║
    ║                                                          By: Nightstalker                ║
    ║              ╔══════════════════════════════════════════════════╗                        ║
    ║              ║             800-171 Rev 2                        ║                        ║
    ║              ╚══════════════════════════════════════════════════╝                        ║
    ║                                                                                          ║
    ║           NIST SP 800-171 Rev 2 Compliance Scanner & Assessment Tool                     ║
    ║                                                                                          ║
    ║   ┌─────────────────────────────────────────────────────────────────────────────────┐    ║
    ║   │  * Automated Network Discovery & Port Scanning                                  │    ║
    ║   │  * NIST SP 800-171 Rev 2 Compliance Assessment                                  │    ║
    ║   │  * SPRS Score Calculation & Risk Analysis                                       │    ║
    ║   │  * Network Topology Visualization                                               │    ║
    ║   │  * Multi-Format Reporting (HTML, JSON, Excel, Text)                             │    ║
    ║   │  * Plan of Action & Milestones (POA&M) Generation                               │    ║
    ║   └─────────────────────────────────────────────────────────────────────────────────┘    ║
    ║                                                                                          ║
    ║               Version: 1.2.0  |  License: MIT  |  Windows Compatible                     ║
    ║                                                                                          ║
    ╚══════════════════════════════════════════════════════════════════════════════════════════╝

    """
    print(banner)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('nistify800-171r2.log', encoding='utf-8'),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)

@dataclass
class ComplianceResult:
    """Result of a compliance check"""
    control_id: str
    control_name: str
    control_text: str
    status: str
    finding: str
    remediation: str
    severity: str
    evidence: List[str]

@dataclass
class NetworkTopology:
    """Network topology information"""
    nodes: Dict[str, Dict]
    edges: List[Tuple[str, str]]
    subnets: List[str]
    gateways: List[str]
    network_diagram_path: Optional[str] = None

@dataclass
class SystemInfo:
    """System information for scanned endpoint"""
    hostname: str
    ip_address: str
    os_type: str
    os_version: str
    open_ports: List[int]
    services: Dict[int, str]
    last_scanned: str
    mac_address: Optional[str] = None
    vendor: Optional[str] = None
    hop_count: Optional[int] = None
    gateway: Optional[str] = None

class NIST80171Controls:
    """NIST SP 800-171 Rev 2 Control Definitions"""
    
    CONTROLS = {
        "3.1.1": {
            "name": "Access Control Policy and Procedures",
            "text": "Establish and maintain baseline configurations and inventories of organizational systems.",
            "family": "AC"
        },
        "3.1.2": {
            "name": "Account Management",
            "text": "Limit system access to authorized users, processes acting on behalf of authorized users, and devices.",
            "family": "AC"
        },
        "3.4.7": {
            "name": "Nonessential Programs",
            "text": "Restrict, disable, or prevent the use of nonessential programs, functions, ports, protocols, and services.",
            "family": "CM"
        },
        "3.13.1": {
            "name": "Boundary Protection",
            "text": "Monitor, control, and protect organizational communications at external and internal boundaries.",
            "family": "SC"
        }
    }

class NetworkScanner:
    """Network and endpoint scanner using nmap"""
    
    def __init__(self):
        try:
            self.nm = nmap.PortScanner()
            self.topology = None
        except nmap.PortScannerError as e:
            logger.error(f"Nmap initialization failed: {e}")
            logger.error("Please install nmap:")
            logger.error("Windows: Download from https://nmap.org/download.html")
            logger.error("Linux: sudo apt-get install nmap")
            logger.error("macOS: brew install nmap")
            raise SystemExit("Nmap is required but not found. Please install nmap and try again.")
        
    def scan_network(self, network_range: str, ports: str = "22,23,53,80,135,139,443,445,993,995") -> List[SystemInfo]:
        """Scan network range for active hosts and services"""
        logger.info(f"[SCAN] Scanning network range: {network_range}")
        systems = []
        
        try:
            is_windows = platform.system().lower() == 'windows'
            
            if is_windows:
                scan_args = '-sn'
            else:
                scan_args = '-sn -PR -PS21,22,23,25,53,80,110,111,135,139,143,443,993,995'
            
            self.nm.scan(hosts=network_range, arguments=scan_args)
            active_hosts = [host for host in self.nm.all_hosts() if self.nm[host].state() == 'up']
            
            logger.info(f"[SCAN] Found {len(active_hosts)} active hosts")
            
            for host in active_hosts:
                try:
                    if is_windows:
                        self.nm.scan(host, ports, arguments='-sV -sS')
                    else:
                        self.nm.scan(host, ports, arguments='-sV -O -A --version-all')
                    
                    if host in self.nm.all_hosts():
                        host_info = self.nm[host]
                        
                        os_type = "Unknown"
                        os_version = "Unknown"
                        
                        if 'osmatch' in host_info and host_info['osmatch']:
                            os_match = host_info['osmatch'][0]
                            os_type = os_match.get('name', 'Unknown')
                            if 'osclass' in os_match and os_match['osclass']:
                                os_version = os_match['osclass'][0].get('osfamily', 'Unknown')
                        
                        open_ports = []
                        services = {}
                        
                        if 'tcp' in host_info:
                            for port, port_info in host_info['tcp'].items():
                                if port_info['state'] == 'open':
                                    open_ports.append(port)
                                    service_name = port_info.get('name', 'unknown')
                                    service_version = port_info.get('version', '')
                                    services[port] = f"{service_name} {service_version}".strip()
                        
                        hostname = host_info.hostname() if host_info.hostname() else host
                        mac_address = None
                        vendor = None
                        
                        if 'addresses' in host_info:
                            addresses = host_info['addresses']
                            if 'mac' in addresses:
                                mac_address = addresses['mac']
                        
                        if 'vendor' in host_info and host_info['vendor']:
                            vendor = list(host_info['vendor'].values())[0] if host_info['vendor'] else None
                        
                        hop_count = None
                        gateway = None
                        if not is_windows and 'traceroute' in host_info:
                            traceroute = host_info['traceroute']
                            if traceroute:
                                hop_count = len(traceroute)
                                if len(traceroute) > 0:
                                    gateway = traceroute[0].get('ipaddr')
                        else:
                            gateway = self._get_default_gateway_windows() if is_windows else None
                        
                        system_info = SystemInfo(
                            hostname=hostname,
                            ip_address=host,
                            os_type=os_type,
                            os_version=os_version,
                            open_ports=open_ports,
                            services=services,
                            last_scanned=datetime.datetime.now().isoformat(),
                            mac_address=mac_address,
                            vendor=vendor,
                            hop_count=hop_count,
                            gateway=gateway
                        )
                        
                        systems.append(system_info)
                        logger.info(f"[SCAN] Scanned {host}: {len(open_ports)} open ports, OS: {os_type}")
                        
                except Exception as e:
                    logger.error(f"[ERROR] Error scanning host {host}: {e}")
                    continue
                    
        except Exception as e:
            logger.error(f"[ERROR] Error during network scan: {e}")
            
        return systems
    
    def _get_default_gateway_windows(self) -> Optional[str]:
        """Get default gateway on Windows systems"""
        try:
            result = subprocess.run(['ipconfig'], capture_output=True, text=True, shell=True)
            lines = result.stdout.split('\n')
            for line in lines:
                if 'Default Gateway' in line and ':' in line:
                    gateway = line.split(':')[1].strip()
                    if gateway and gateway != '':
                        return gateway
        except Exception as e:
            logger.debug(f"Could not determine default gateway: {e}")
        return None
    
    def discover_topology(self, systems: List[SystemInfo]) -> NetworkTopology:
        """Discover and map network topology from scan results"""
        logger.info("[TOPOLOGY] Analyzing network topology...")
        
        nodes = {}
        edges = []
        subnets = set()
        gateways = set()
        
        for system in systems:
            node_info = {
                'hostname': system.hostname,
                'ip': system.ip_address,
                'os_type': system.os_type,
                'open_ports': len(system.open_ports),
                'services': list(system.services.values())[:3],
                'mac_address': system.mac_address,
                'vendor': system.vendor,
                'hop_count': system.hop_count or 1,
                'type': self._classify_node_type(system)
            }
            nodes[system.ip_address] = node_info
            
            try:
                network = ipaddress.ip_network(f"{system.ip_address}/24", strict=False)
                subnets.add(str(network))
            except:
                pass
            
            if system.gateway:
                gateways.add(system.gateway)
                if system.gateway != system.ip_address:
                    edges.append((system.ip_address, system.gateway))
        
        for gateway in gateways:
            if gateway not in nodes:
                nodes[gateway] = {
                    'hostname': f'Gateway-{gateway}',
                    'ip': gateway,
                    'os_type': 'Gateway/Router',
                    'open_ports': 0,
                    'services': ['Routing'],
                    'mac_address': None,
                    'vendor': 'Unknown',
                    'hop_count': 0,
                    'type': 'gateway'
                }
        
        topology = NetworkTopology(
            nodes=nodes,
            edges=edges,
            subnets=list(subnets),
            gateways=list(gateways)
        )
        
        self.topology = topology
        return topology
    
    def _classify_node_type(self, system: SystemInfo) -> str:
        """Classify node type based on services and characteristics"""
        services = [service.lower() for service in system.services.values()]
        open_ports = system.open_ports
        
        if any('http' in service or 'web' in service for service in services):
            return 'web_server'
        elif 22 in open_ports or any('ssh' in service for service in services):
            return 'server'
        elif 'windows' in system.os_type.lower():
            return 'windows_client'
        elif 'linux' in system.os_type.lower():
            return 'linux_client'
        
        return 'unknown'
    
    def create_network_diagram(self, topology: NetworkTopology, output_path: str = "network_topology.png"):
        """Create a visual network topology diagram"""
        logger.info("[DIAGRAM] Generating network topology diagram...")
        
        try:
            G = nx.Graph()
            
            for ip, node_info in topology.nodes.items():
                G.add_node(ip, **node_info)
            
            G.add_edges_from(topology.edges)
            
            plt.figure(figsize=(16, 12))
            plt.clf()
            
            node_colors = {
                'gateway': '#FF6B6B',
                'web_server': '#4ECDC4',
                'server': '#45B7D1',
                'windows_client': '#AED6F1',
                'linux_client': '#A9DFBF',
                'unknown': '#D5DBDB'
            }
            
            pos = nx.spring_layout(G, k=3, iterations=50, seed=42)
            
            for node_type, color in node_colors.items():
                nodes_of_type = [node for node, data in G.nodes(data=True) if data.get('type') == node_type]
                if nodes_of_type:
                    node_sizes = [1000 + (G.nodes[node].get('open_ports', 0) * 100) for node in nodes_of_type]
                    nx.draw_networkx_nodes(G, pos, nodelist=nodes_of_type, 
                                         node_color=color, node_size=node_sizes, 
                                         alpha=0.8, edgecolors='black', linewidths=1)
            
            nx.draw_networkx_edges(G, pos, alpha=0.6)
            
            labels = {}
            for node, data in G.nodes(data=True):
                hostname = data.get('hostname', node)
                if hostname != node:
                    labels[node] = f"{hostname}\n{node}"
                else:
                    labels[node] = node
            
            nx.draw_networkx_labels(G, pos, labels, font_size=8, font_weight='bold')
            
            legend_elements = []
            for node_type, color in node_colors.items():
                if any(data.get('type') == node_type for _, data in G.nodes(data=True)):
                    legend_elements.append(mpatches.Patch(color=color, label=node_type.replace('_', ' ').title()))
            
            plt.legend(handles=legend_elements, loc='upper left', bbox_to_anchor=(0, 1))
            
            plt.title("Network Topology Diagram\nNISTify 800-171 R2 Compliance Assessment", 
                     fontsize=16, fontweight='bold', pad=20)
            
            info_text = f"Total Nodes: {len(G.nodes())}\nTotal Connections: {len(G.edges())}\n"
            info_text += f"Subnets: {len(topology.subnets)}\nGateways: {len(topology.gateways)}"
            
            plt.text(0.02, 0.98, info_text, transform=plt.gca().transAxes, 
                    verticalalignment='top', bbox=dict(boxstyle='round', facecolor='wheat', alpha=0.8))
            
            plt.axis('off')
            plt.tight_layout()
            plt.savefig(output_path, dpi=300, bbox_inches='tight', 
                       facecolor='white', edgecolor='none')
            plt.close()
            
            topology.network_diagram_path = output_path
            logger.info(f"[DIAGRAM] Network topology diagram saved: {output_path}")
            
            return output_path
            
        except Exception as e:
            logger.error(f"[ERROR] Error creating network diagram: {e}")
            return None

class ComplianceAssessor:
    """Assess NIST SP 800-171 compliance based on scan results"""
    
    def __init__(self):
        self.controls = NIST80171Controls.CONTROLS
        
    def assess_system(self, system: SystemInfo) -> List[ComplianceResult]:
        """Assess a single system for compliance"""
        results = []
        
        logger.info(f"[ASSESS] Assessing compliance for {system.hostname} ({system.ip_address})")
        
        weak_services = self._check_weak_services(system)
        if weak_services:
            results.append(ComplianceResult(
                control_id="3.1.2",
                control_name=self.controls["3.1.2"]["name"],
                control_text=self.controls["3.1.2"]["text"],
                status="non_compliant",
                finding=f"Potentially insecure services detected: {', '.join(weak_services)}",
                remediation="Disable unnecessary services, implement strong authentication, and restrict access",
                severity="high",
                evidence=[f"Open ports: {system.open_ports}", f"Services: {system.services}"]
            ))
        else:
            results.append(ComplianceResult(
                control_id="3.1.2",
                control_name=self.controls["3.1.2"]["name"],
                control_text=self.controls["3.1.2"]["text"],
                status="compliant",
                finding="No obviously insecure services detected",
                remediation="Continue monitoring for unauthorized services",
                severity="low",
                evidence=[f"Services reviewed: {list(system.services.values())}"]
            ))
        
        unnecessary_ports = self._check_unnecessary_ports(system)
        if unnecessary_ports:
            results.append(ComplianceResult(
                control_id="3.4.7",
                control_name=self.controls["3.4.7"]["name"],
                control_text=self.controls["3.4.7"]["text"],
                status="non_compliant",
                finding=f"Potentially unnecessary ports open: {unnecessary_ports}",
                remediation="Review and close unnecessary ports, disable unused services",
                severity="medium",
                evidence=[f"Open ports: {system.open_ports}"]
            ))
        
        external_services = self._check_external_services(system)
        if external_services:
            results.append(ComplianceResult(
                control_id="3.13.1",
                control_name=self.controls["3.13.1"]["name"],
                control_text=self.controls["3.13.1"]["text"],
                status="non_compliant",
                finding=f"External-facing services detected: {external_services}",
                remediation="Implement firewall rules, access controls, and monitoring for external-facing services",
                severity="high",
                evidence=[f"External services: {external_services}"]
            ))
        
        return results
    
    def _check_weak_services(self, system: SystemInfo) -> List[str]:
        """Check for potentially weak or insecure services"""
        weak_services = []
        risky_ports = {21: "FTP", 23: "Telnet", 135: "RPC", 139: "NetBIOS", 445: "SMB"}
        
        for port in system.open_ports:
            if port in risky_ports:
                service_name = system.services.get(port, risky_ports[port])
                weak_services.append(f"{service_name} (port {port})")
                
        return weak_services
    
    def _check_unnecessary_ports(self, system: SystemInfo) -> List[int]:
        """Check for potentially unnecessary open ports"""
        essential_ports = {22, 80, 443}
        return [port for port in system.open_ports if port not in essential_ports]
    
    def _check_external_services(self, system: SystemInfo) -> List[str]:
        """Check for services that might be externally accessible"""
        external_services = []
        external_ports = {21, 22, 23, 80, 443, 993, 995}
        
        for port in system.open_ports:
            if port in external_ports:
                service_name = system.services.get(port, f"Port {port}")
                external_services.append(service_name)
                
        return external_services

class SPRSCalculator:
    """Calculate SPRS (Supplier Performance Risk System) score"""
    
    def calculate_sprs_score(self, results: List[ComplianceResult]) -> Dict:
        """Calculate SPRS score based on compliance results"""
        logger.info("[SPRS] Calculating SPRS compliance score...")
        
        total_controls = len(self.get_all_control_ids())
        
        compliant = len([r for r in results if r.status == 'compliant'])
        non_compliant = len([r for r in results if r.status == 'non_compliant'])
        not_applicable = len([r for r in results if r.status == 'not_applicable'])
        not_assessed = len([r for r in results if r.status == 'not_assessed'])
        
        applicable_controls = total_controls - not_applicable
        if applicable_controls > 0:
            compliance_percentage = (compliant / applicable_controls) * 100
        else:
            compliance_percentage = 100
            
        base_score = 110
        
        high_severity_deduction = len([r for r in results if r.status == 'non_compliant' and r.severity == 'high']) * 15
        medium_severity_deduction = len([r for r in results if r.status == 'non_compliant' and r.severity == 'medium']) * 10
        low_severity_deduction = len([r for r in results if r.status == 'non_compliant' and r.severity == 'low']) * 5
        
        total_deduction = high_severity_deduction + medium_severity_deduction + low_severity_deduction
        sprs_score = max(0, base_score - total_deduction)
        
        return {
            'sprs_score': sprs_score,
            'max_score': base_score,
            'compliance_percentage': round(compliance_percentage, 2),
            'total_controls': total_controls,
            'compliant': compliant,
            'non_compliant': non_compliant,
            'not_applicable': not_applicable,
            'not_assessed': not_assessed,
            'high_severity_findings': len([r for r in results if r.status == 'non_compliant' and r.severity == 'high']),
            'medium_severity_findings': len([r for r in results if r.status == 'non_compliant' and r.severity == 'medium']),
            'low_severity_findings': len([r for r in results if r.status == 'non_compliant' and r.severity == 'low'])
        }
    
    def get_all_control_ids(self) -> List[str]:
        """Get all NIST SP 800-171 control IDs"""
        return list(NIST80171Controls.CONTROLS.keys())

class ReportGenerator:
    """Generate compliance reports in multiple formats"""
    
    def __init__(self, output_dir: str = "reports"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True)
        
    def generate_all_reports(self, systems: List[SystemInfo], results: List[ComplianceResult], sprs_data: Dict, topology: NetworkTopology = None):
        """Generate reports in all formats"""
        logger.info("[REPORTS] Generating compliance reports in multiple formats...")
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        
        self.generate_html_report(systems, results, sprs_data, f"nistify_compliance_report_{timestamp}.html", topology)
        self.generate_json_report(systems, results, sprs_data, f"nistify_compliance_report_{timestamp}.json", topology)
        self.generate_text_report(systems, results, sprs_data, f"nistify_compliance_report_{timestamp}.txt")
        self.generate_poam_xlsx(results, f"nistify_poam_{timestamp}.xlsx")
        
        logger.info(f"[REPORTS] Reports generated in {self.output_dir}")
    
    def generate_html_report(self, systems: List[SystemInfo], results: List[ComplianceResult], sprs_data: Dict, filename: str, topology: NetworkTopology = None):
        """Generate HTML compliance report"""
        html_content = f"""
<!DOCTYPE html>
<html>
<head>
    <title>NISTify 800-171 R2 - Compliance Report</title>
    <style>
        body {{ font-family: Arial, sans-serif; margin: 20px; background-color: #f8f9fa; }}
        .header {{ background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; padding: 30px; border-radius: 10px; text-align: center; }}
        .summary {{ background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; padding: 25px; margin: 20px 0; border-radius: 10px; }}
        .sprs-score {{ font-size: 36px; font-weight: bold; text-align: center; background: rgba(255,255,255,0.2); padding: 20px; border-radius: 10px; margin: 15px 0; }}
        table {{ border-collapse: collapse; width: 100%; margin: 20px 0; background: white; border-radius: 10px; overflow: hidden; }}
        th, td {{ border: none; padding: 12px 15px; text-align: left; }}
        th {{ background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; font-weight: bold; }}
        tr:nth-child(even) {{ background-color: #f8f9fa; }}
        .compliant {{ background-color: #d4edda !important; }}
        .non-compliant {{ background-color: #f8d7da !important; }}
        .high-severity {{ color: #dc3545; font-weight: bold; }}
        .medium-severity {{ color: #fd7e14; font-weight: bold; }}
        .low-severity {{ color: #28a745; font-weight: bold; }}
    </style>
</head>
<body>
    <div class="header">
        <h1>NISTify 800-171 R2</h1>
        <p>NIST SP 800-171 Rev 2 Compliance Assessment Report</p>
        <p>Generated on: {datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")}</p>
    </div>
    
    <div class="summary">
        <h2>Executive Summary</h2>
        <div class="sprs-score">
            SPRS Score: {sprs_data['sprs_score']} / {sprs_data['max_score']}
            <div style="font-size: 18px; margin-top: 10px;">
                Compliance: {sprs_data['compliance_percentage']}%
            </div>
        </div>
        <p>Systems Assessed: {len(systems)}</p>
        <p>Total Findings: {len(results)}</p>
        <p>High Severity: {sprs_data['high_severity_findings']}, Medium: {sprs_data['medium_severity_findings']}, Low: {sprs_data['low_severity_findings']}</p>
    </div>
    
    <h2>Scanned Systems</h2>
    <table>
        <tr>
            <th>Hostname</th>
            <th>IP Address</th>
            <th>OS Type</th>
            <th>Open Ports</th>
            <th>Last Scanned</th>
        </tr>
"""
        
        for system in systems:
            html_content += f"""
        <tr>
            <td><strong>{system.hostname}</strong></td>
            <td>{system.ip_address}</td>
            <td>{system.os_type}</td>
            <td>{', '.join(map(str, system.open_ports))}</td>
            <td>{system.last_scanned}</td>
        </tr>
"""
        
        html_content += """
    </table>
    
    <h2>Compliance Findings</h2>
    <table>
        <tr>
            <th>Control ID</th>
            <th>Control Name</th>
            <th>Status</th>
            <th>Severity</th>
            <th>Finding</th>
            <th>Remediation</th>
        </tr>
"""
        
        for result in results:
            status_class = result.status.replace('_', '-')
            severity_class = f"{result.severity}-severity"
            html_content += f"""
        <tr class="{status_class}">
            <td><strong>{result.control_id}</strong></td>
            <td>{result.control_name}</td>
            <td>{result.status.replace('_', ' ').title()}</td>
            <td class="{severity_class}">{result.severity.title()}</td>
            <td>{result.finding}</td>
            <td>{result.remediation}</td>
        </tr>
"""
        
        html_content += """
    </table>
    
    <div style="text-align: center; margin-top: 40px; padding: 20px; background: #f8f9fa; border-radius: 10px;">
        <p style="color: #666; margin: 0;">
            Generated by NISTify 800-171 R2 v1.2.0 | 
            Comprehensive NIST SP 800-171 Rev 2 Compliance Assessment Tool
        </p>
    </div>
</body>
</html>
        """
        
        with open(self.output_dir / filename, 'w', encoding='utf-8') as f:
            f.write(html_content)
    
    def generate_json_report(self, systems: List[SystemInfo], results: List[ComplianceResult], sprs_data: Dict, filename: str, topology: NetworkTopology = None):
        """Generate JSON compliance report"""
        report_data = {
            "metadata": {
                "generated_on": datetime.datetime.now().isoformat(),
                "standard": "NIST SP 800-171 Rev 2",
                "tool": "NISTify 800-171 R2",
                "version": "1.2.0"
            },
            "sprs_score": sprs_data,
            "scanned_systems": [asdict(system) for system in systems],
            "compliance_results": [asdict(result) for result in results]
        }
        
        if topology:
            report_data["network_topology"] = {
                "nodes": topology.nodes,
                "edges": topology.edges,
                "subnets": topology.subnets,
                "gateways": topology.gateways,
                "diagram_path": topology.network_diagram_path
            }
        
        with open(self.output_dir / filename, 'w', encoding='utf-8') as f:
            json.dump(report_data, f, indent=2, ensure_ascii=False)
    
    def generate_text_report(self, systems: List[SystemInfo], results: List[ComplianceResult], sprs_data: Dict, filename: str):
        """Generate text compliance report"""
        with open(self.output_dir / filename, 'w', encoding='utf-8') as f:
            f.write("╔══════════════════════════════════════════════════════════════════════════════════════════════╗\n")
            f.write("║                          NISTify 800-171 R2 COMPLIANCE REPORT                               ║\n")
            f.write("╚══════════════════════════════════════════════════════════════════════════════════════════════╝\n\n")
            
            f.write(f"Generated on: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"Standard: NIST SP 800-171 Rev 2\n")
            f.write(f"Tool: NISTify 800-171 R2 v1.2.0\n\n")
            
            f.write("EXECUTIVE SUMMARY\n")
            f.write("=" * 50 + "\n")
            f.write(f"SPRS Score: {sprs_data['sprs_score']} / {sprs_data['max_score']}\n")
            f.write(f"Compliance Percentage: {sprs_data['compliance_percentage']}%\n")
            f.write(f"Systems Assessed: {len(systems)}\n")
            f.write(f"Total Findings: {len(results)}\n")
            f.write(f"High Severity: {sprs_data['high_severity_findings']}\n")
            f.write(f"Medium Severity: {sprs_data['medium_severity_findings']}\n")
            f.write(f"Low Severity: {sprs_data['low_severity_findings']}\n\n")
            
            f.write("SCANNED SYSTEMS\n")
            f.write("=" * 30 + "\n")
            for system in systems:
                f.write(f"Hostname: {system.hostname}\n")
                f.write(f"IP Address: {system.ip_address}\n")
                f.write(f"OS Type: {system.os_type}\n")
                f.write(f"Open Ports: {', '.join(map(str, system.open_ports))}\n")
                f.write(f"Last Scanned: {system.last_scanned}\n\n")
            
            f.write("COMPLIANCE FINDINGS\n")
            f.write("=" * 35 + "\n")
            for result in results:
                f.write(f"Control ID: {result.control_id}\n")
                f.write(f"Control Name: {result.control_name}\n")
                f.write(f"Status: {result.status.replace('_', ' ').title()}\n")
                f.write(f"Severity: {result.severity.title()}\n")
                f.write(f"Finding: {result.finding}\n")
                f.write(f"Remediation: {result.remediation}\n")
                if result.evidence:
                    f.write(f"Evidence: {'; '.join(result.evidence)}\n")
                f.write("\n" + "-" * 80 + "\n\n")
            
            f.write("\n" + "="*80 + "\n")
            f.write("Generated by NISTify 800-171 R2 v1.2.0\n")
            f.write("Comprehensive NIST SP 800-171 Rev 2 Compliance Assessment Tool\n")
            f.write("="*80 + "\n")
    
    def generate_poam_xlsx(self, results: List[ComplianceResult], filename: str):
        """Generate Plan of Action and Milestones (POA&M) Excel document"""
        logger.info("[POAM] Generating POA&M Excel document...")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "NISTify POA&M"
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
        high_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
        medium_fill = PatternFill(start_color="FFD93D", end_color="FFD93D", fill_type="solid")
        low_fill = PatternFill(start_color="6BCF7F", end_color="6BCF7F", fill_type="solid")
        
        headers = [
            "Control Number", "Control Name", "Control Text", "Status", "Severity",
            "Deficiency Identified", "Remediation Steps", "Target Date", 
            "Responsible Party", "Status Notes", "Evidence"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        row = 2
        for result in results:
            if result.status == 'non_compliant':
                ws.cell(row, 1, result.control_id)
                ws.cell(row, 2, result.control_name)
                ws.cell(row, 3, result.control_text)
                ws.cell(row, 4, result.status.replace('_', ' ').title())
                
                severity_cell = ws.cell(row, 5, result.severity.title())
                if result.severity == 'high':
                    severity_cell.fill = high_fill
                elif result.severity == 'medium':
                    severity_cell.fill = medium_fill
                else:
                    severity_cell.fill = low_fill
                
                ws.cell(row, 6, result.finding)
                ws.cell(row, 7, result.remediation)
                
                target_date = datetime.datetime.now()
                if result.severity == 'high':
                    target_date += datetime.timedelta(days=30)
                elif result.severity == 'medium':
                    target_date += datetime.timedelta(days=90)
                else:
                    target_date += datetime.timedelta(days=180)
                
                ws.cell(row, 8, target_date.strftime("%Y-%m-%d"))
                ws.cell(row, 9, "IT Security Team")
                ws.cell(row, 10, "Open")
                ws.cell(row, 11, '; '.join(result.evidence) if result.evidence else "")
                
                row += 1
        
        column_widths = [15, 30, 50, 15, 10, 40, 50, 12, 20, 15, 30]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[ws.cell(1, col).column_letter].width = width
        
        ws.auto_filter.ref = f"A1:{ws.cell(1, len(headers)).coordinate}"
        
        wb.save(self.output_dir / filename)
        logger.info(f"[POAM] POA&M Excel document generated: {filename}")

class ComplianceScanner:
    """Main NISTify compliance scanner orchestrator"""
    
    def __init__(self, output_dir: str = "reports"):
        self.scanner = NetworkScanner()
        self.assessor = ComplianceAssessor()
        self.sprs_calculator = SPRSCalculator()
        self.report_generator = ReportGenerator(output_dir)
        
    def scan_and_assess(self, network_ranges: List[str], ports: str = None, generate_topology: bool = True) -> Tuple[List[SystemInfo], List[ComplianceResult], Dict, NetworkTopology]:
        """Perform complete scan and assessment"""
        logger.info("[START] Starting NISTify 800-171 R2 compliance assessment...")
        
        all_systems = []
        all_results = []
        topology = None
        
        for network_range in network_ranges:
            logger.info(f"[SCAN] Scanning network range: {network_range}")
            systems = self.scanner.scan_network(network_range, ports)
            all_systems.extend(systems)
        
        if generate_topology and all_systems:
            topology = self.scanner.discover_topology(all_systems)
            diagram_path = str(Path(self.report_generator.output_dir) / "nistify_network_topology.png")
            self.scanner.create_network_diagram(topology, diagram_path)
        
        logger.info(f"[ASSESS] Assessing {len(all_systems)} systems for NIST SP 800-171 compliance")
        for system in all_systems:
            results = self.assessor.assess_system(system)
            all_results.extend(results)
        
        sprs_data = self.sprs_calculator.calculate_sprs_score(all_results)
        
        return all_systems, all_results, sprs_data, topology
    
    def generate_reports(self, systems: List[SystemInfo], results: List[ComplianceResult], sprs_data: Dict, topology: NetworkTopology = None):
        """Generate all compliance reports"""
        self.report_generator.generate_all_reports(systems, results, sprs_data, topology)

def main():
    """Main entry point"""
    print_banner()
    
    parser = argparse.ArgumentParser(
        description="NISTify 800-171 R2 - NIST SP 800-171 Rev 2 Compliance Scanner",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python nistify800-171r2.py 192.168.1.0/24
  python nistify800-171r2.py 10.0.0.0/8 192.168.0.0/16 --verbose
  python nistify800-171r2.py 172.16.0.0/12 --ports "22,80,443,3389" --no-topology
        """
    )
    
    parser.add_argument("networks", nargs="+", 
                       help="Network ranges to scan (e.g., 192.168.1.0/24)")
    parser.add_argument("--ports", default="22,23,53,80,135,139,443,445,993,995", 
                       help="Comma-separated ports to scan (default: common ports)")
    parser.add_argument("--output-dir", default="nistify_reports", 
                       help="Output directory for reports (default: nistify_reports)")
    parser.add_argument("--no-topology", action="store_true", 
                       help="Skip network topology generation for faster scanning")
    parser.add_argument("--verbose", "-v", action="store_true", help="Enable verbose logging")
    
    args = parser.parse_args()
    
    if args.verbose:
        logging.getLogger().setLevel(logging.DEBUG)
        logger.info("[CONFIG] Verbose logging enabled")
    
    valid_networks = []
    logger.info("[CONFIG] Validating network ranges...")
    for network in args.networks:
        try:
            ipaddress.ip_network(network, strict=False)
            valid_networks.append(network)
            logger.info(f"[CONFIG] Valid network range: {network}")
        except ValueError:
            logger.error(f"[ERROR] Invalid network range: {network}")
            continue
    
    if not valid_networks:
        logger.error("[ERROR] No valid network ranges provided")
        sys.exit(1)
    
    scanner = ComplianceScanner(args.output_dir)
    
    try:
        logger.info("[START] Starting NISTify 800-171 R2 compliance assessment")
        start_time = datetime.datetime.now()
        
        systems, results, sprs_data, topology = scanner.scan_and_assess(
            valid_networks, args.ports, not args.no_topology
        )
        
        end_time = datetime.datetime.now()
        duration = end_time - start_time
        
        logger.info(f"[COMPLETE] Assessment complete in {duration.total_seconds():.1f} seconds")
        logger.info(f"[RESULTS] Found {len(systems)} systems with {len(results)} findings")
        logger.info(f"[SPRS] SPRS Score: {sprs_data['sprs_score']} / {sprs_data['max_score']}")
        
        if topology:
            logger.info(f"[TOPOLOGY] Network topology: {len(topology.nodes)} nodes, {len(topology.edges)} connections")
            logger.info(f"[TOPOLOGY] Discovered subnets: {', '.join(topology.subnets)}")
        
        scanner.generate_reports(systems, results, sprs_data, topology)
        
        print(f"\n{'='*90}")
        print("NISTify 800-171 R2 COMPLIANCE ASSESSMENT COMPLETE")
        print(f"{'='*90}")
        print(f"Assessment Duration: {duration.total_seconds():.1f} seconds")
        print(f"Systems Scanned: {len(systems)}")
        print(f"Compliance Findings: {len(results)}")
        print(f"SPRS Score: {sprs_data['sprs_score']} / {sprs_data['max_score']}")
        print(f"Compliance Rate: {sprs_data['compliance_percentage']}%")
        print(f"High Severity Issues: {sprs_data['high_severity_findings']}")
        print(f"Medium Severity Issues: {sprs_data['medium_severity_findings']}")
        print(f"Low Severity Issues: {sprs_data['low_severity_findings']}")
        
        if topology:
            print(f"\nNetwork Topology Analysis:")
            print(f"   Total Nodes: {len(topology.nodes)}")
            print(f"   Network Connections: {len(topology.edges)}")
            print(f"   Subnets Discovered: {len(topology.subnets)}")
            print(f"   Gateways Identified: {len(topology.gateways)}")
        
        print(f"\nReports generated in: {args.output_dir}")
        print("   HTML Report: nistify_compliance_report_*.html")
        print("   JSON Report: nistify_compliance_report_*.json")
        print("   Text Report: nistify_compliance_report_*.txt")
        print("   POA&M Document: nistify_poam_*.xlsx")
        
        if topology and topology.network_diagram_path:
            print("   Network Topology Diagram: nistify_network_topology.png")
        
        print(f"\nThank you for using NISTify 800-171 R2!")
        print("   For support and updates: https://github.com/yourusername/nistify800-171r2")
        
    except KeyboardInterrupt:
        logger.info("\n[INTERRUPTED] Assessment interrupted by user")
        sys.exit(0)
    except Exception as e:
        logger.error(f"[ERROR] Assessment failed: {e}")
        if args.verbose:
            import traceback
            logger.error(f"Full traceback: {traceback.format_exc()}")
        sys.exit(1)

if __name__ == "__main__":
    main()